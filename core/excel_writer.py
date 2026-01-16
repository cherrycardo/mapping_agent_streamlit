from openpyxl import load_workbook

def _norm(s) -> str:
    return " ".join(str(s or "").strip().lower().split())

def _find_cell(ws, needle: str, max_rows: int = 50):
    needle_n = _norm(needle)
    for r in range(1, min(max_rows, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            if _norm(ws.cell(r, c).value) == needle_n:
                return r, c
    return None, None

def append_raw_bronze_to_template(
    template_path: str,
    output_path: str,
    sheet_name: str,
    raw_table_name: str,
    bronze_table_name: str,
    pairs: list[dict],
):
    wb = load_workbook(template_path)
    ws = wb[sheet_name]

    # Find header row by locating the "Raw table name" cell
    header_row, _ = _find_cell(ws, "Raw table name", max_rows=50)
    if not header_row:
        raise ValueError(f"Could not find 'Raw table name' header in sheet: {sheet_name}")

    # Build a header -> column index map from that row
    header_to_col = {}
    for c in range(1, ws.max_column + 1):
        h = _norm(ws.cell(header_row, c).value)
        if h:
            header_to_col[h] = c

   def col_any(*header_options: str) -> int:
    for h in header_options:
        key = _norm(h)
        if key in header_to_col:
            return header_to_col[key]
    raise ValueError(f"None of these headers exist in template: {header_options}")
            )
        return header_to_col[key]

    # Find first empty row using Raw column name as the anchor
    r = header_row + 1
    while ws.cell(r, col("Raw column name")).value not in (None, ""):
        r += 1

    for p in pairs:
        # RAW section
        ws.cell(r, col("Raw table name")).value = raw_table_name
        ws.cell(r, col("Raw column name")).value = p.get("raw_column", "")

        # BRONZE section (in your template these are labeled as Table Name / Column Name)
        ws.cell(r, col_any("Bronze Table Name", "Table Name")).value = bronze_table_name
        ws.cell(r, col_any("Bronze Column Name", "Column Name")).value = p.get("bronze_column", "")


        # Optional: if you want to fill Bronze datatype + description in MVP:
        # Bronze datatype header in your template is also "Data Type w/ Precision"
        # Bronze description header is "Column Definition"
        if "bronze_datatype" in p and _norm("Data Type w/ Precision") in header_to_col:
            ws.cell(r, col("Data Type w/ Precision")).value = p.get("bronze_datatype", "")

        if "bronze_description" in p and _norm("Column Definition") in header_to_col:
            ws.cell(r, col("Column Definition")).value = p.get("bronze_description", "")

        r += 1

    wb.save(output_path)

